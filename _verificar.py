import requests         # pip install requests
from pandas import DataFrame , read_excel, read_excel, concat  # pip install pandas
import numpy as np      # pip install numpy
from openpyxl import Workbook , load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.drawing.image import Image
import datetime
import os

def tags_tabela():
    planilha = load_workbook('TimesAPI.xlsx')
    paginaPlayers = planilha['Times']
    timesTag = []
    times = []
    nomesTime = []
    for coluna in paginaPlayers.iter_rows(min_col=1,max_col=5,min_row=2):
        time = []
        for celula in coluna:
            if celula.value != None and celula.value != '':
                tag = celula.value.replace('#','')
                time.append(tag)
        timesTag.append(time)
    for i in range(len(timesTag)):
        tagJogador = []
        nomeTime = []
        for p in range(len(timesTag[i])):
            if p == 0:
                nomeTime = [timesTag[i][p]]
                nomesTime.append(nomeTime[0])
            else:
                tag = timesTag[i][p]
                tagJogador.append(tag)
        if nomeTime != [] and tagJogador != []:
            temp = nomeTime + [tagJogador]
            times.append(temp)
    return [times, nomesTime]

def nomear_player(times):
    tagsPlayers = []
    print('Obtendo dados...')
    for i in range(len(times)):
        tags = []
        print()
        print(f'I: {times[i][0]}')
        for p in range(len(times[i][1])):
            player = []
            nome = namePlayer(times[i][1][p])
            tag = times[i][1][p]
            player.append(tag)
            player.append(nome)
            print(player)
            tags.append(player)
        tagsPlayers.append(tags)
    return tagsPlayers

def namePlayer(jogador):
    with open('token.txt', 'r', encoding='utf-8') as arquivo:
        token = arquivo.read().strip()
    url_api = f"https://api.brawlstars.com/v1/players/%23{jogador}"
    headers = {
        "Authorization": f"Bearer {token}"
    }
    response = requests.get(url_api, headers=headers)
    if response.status_code == 200:
        data = response.json()
        data = data['name']
    else:
        data = None
    return data

def verificar_bot(players):
    bot = False
    for element in players:
        if 'Bot' in element:
            bot = True
            return bot

def verificar_arquivo(nome):
    planilha = load_workbook('TimesAPI.xlsx')
    if nome in planilha.sheetnames:

        print('Arquivo ja existe')
    else:
        while True:
            try:
                planilha.create_sheet(nome)
                planilha.save('TimesAPI.xlsx')
                print('Novo arquivo criado')
                break
            except PermissionError:
                print('Feche o arquivo para que as alteracoes sejam feitas')
                repeat = input('Digite P para cancelar: ')
                if repeat in 'pP':
                    break

def atualizarNomes():
    planilha = load_workbook('TimesAPI copy.xlsx')
    paginaPlayers = planilha['Times']
    times = tags_tabela()
    temp = nomear_player(times[0])
    todosPlayers = []
    todosTags = []
    quant = 0
    print()
    for i in range(len(temp)):
        colunaTag = []
        colunaPlayer = []
        quant+=1
        for j in range(len(temp[i])):
            colunaTag.append(temp[i][j][0])
            colunaPlayer.append(temp[i][j][1])
        todosPlayers.append(colunaPlayer)
        todosTags.append(colunaTag)
    cont=1
    for p in range(len(todosTags)):
        tempTag = []
        for l in range(len(todosTags[p])):
            tempTag.append(todosTags[p][l])
        cont+=1
        quantPlayers = len(tempTag) + 1
        for row in paginaPlayers.iter_rows(min_col=2, min_row=cont, max_row=cont, max_col=quantPlayers):
            cont2 = 0
            lock = None 
            for cell in row:
                if tempTag[cont2] != lock:
                    cell.value = tempTag[cont2]
                lock = tempTag[cont2]
                cont2 += 1
    cont=1
    for p in range(len(todosPlayers)):
        tempPlayer = []
        for l in range(len(todosPlayers[p])):
            tempPlayer.append(todosPlayers[p][l])
        cont+=1
        quantPlayers = len(tempPlayer) + 5
        for row in paginaPlayers.iter_rows(min_col=6,min_row=cont,max_row=cont, max_col=quantPlayers):           
            cont2=0
            for cell in row:
                cell.value = tempPlayer[cont2]
                cont2+=1
    print(f'Quantidade de times: {quant}')
    while True:
        try:
            planilha.save('TimesAPI.xlsx')
            print()
            print('*Arquivo atualizado*')
            break
        except PermissionError:
            print('-----------')
            input('ERRO! Feche o arquivo para que as alteracoes sejam feitas. Pressione Enter para tentar novamente: ')

def Salvar(tabelaFinal, nomeArquivo):
    # Verificar se o arquivo já existe
    if os.path.exists(f'{nomeArquivo}.xlsx'):
        # Se o arquivo existir, ler os dados
        dados_salvos = read_excel(f'{nomeArquivo}.xlsx')
        
        # Identificar linhas novas
        linhas_novas = tabelaFinal[~tabelaFinal.apply(tuple, axis=1).isin(dados_salvos.apply(tuple, axis=1))]
        
        # Concatenar linhas novas com dados existentes
        dados = concat([linhas_novas, dados_salvos], ignore_index=True)  # Novas linhas no início
        
        # Se houver linhas novas para adicionar
        if not linhas_novas.empty:
            # Salvando os dados em um arquivo Excel
            while True:
                try:
                    dados.to_excel(f'{nomeArquivo}.xlsx', index=False, engine='openpyxl')
                    print("Dados salvos com sucesso.")
                    break
                except PermissionError:
                    input("Feche o arquivo para o salvamento ser concluído. ")
        else:
            print("Nenhuma linha nova para adicionar. Nenhum arquivo modificado.")
    else:
        # Se o arquivo não existir, salvar 'tabelaFinal' diretamente
        tabelaFinal.to_excel(f'{nomeArquivo}.xlsx', index=False)
        print("Arquivo criado e dados salvos com sucesso.")


def formatarDocumento(nomeArquivo):
    workbook = load_workbook(f'{nomeArquivo}.xlsx')
    sheet = workbook.active

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  
        for cell in sheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    for row in sheet.iter_rows():
        max_height = 0
        for cell in row:
            try:
                cell_text = str(cell.value)
                lines = cell_text.split('\n')
                height = max(len(line) for line in lines)
                if height > max_height:
                    max_height = height
            except:
                pass
        sheet.row_dimensions[row[0].row].height = max_height * 1.5

    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            if cell.value == 'defeat':
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Vermelho
            elif cell.value == 'victory':
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
            elif cell.value == 'draw':
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Cinza

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == 'gemGrab':
                cell.fill = PatternFill(start_color="8a2be2", end_color="8a2be2", fill_type="solid")  # roxo escuro
            elif cell.value == 'bounty':
                cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")  # amarelo
            elif cell.value == 'hotZone':
                cell.fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")  # vermelho
            elif cell.value == 'brawlBall':
                cell.fill = PatternFill(start_color="0000ff", end_color="0000ff", fill_type="solid")  # azul
            elif cell.value == 'knockout':
                cell.fill = PatternFill(start_color="ffa500", end_color="ffa500", fill_type="solid")  # laranja
            elif cell.value == 'heist':
                cell.fill = PatternFill(start_color="d8bfd8", end_color="d8bfd8", fill_type="solid")  # roxo claro

    # Carregar a planilha existente
    

    diretorio_imagens = '/brawlers'

    # Lista de nomes dos Brawlers
    nomes_brawlers = [
        'Angelo', 'Melodie', 'Larry & Lawrie', 'Kit', 'Mico', 'Charlie', 'Chuck', 'Pearl', 'Doug', 'Cordelius',
        'Hank', 'Maisie', 'Willow', 'R-T', 'Mandy', 'Gray', 'Chester', 'Buster', 'Gus', 'Sam', 'Otis', 'Bonnie',
        'Janet', 'Eve', 'Fang', 'Lolla', 'Meg', 'Ash', 'Griff', 'Buzz', 'Grom', 'Squeak', 'Belle', 'Stu', 'Ruffs',
        'Edgar', 'Byron', 'Lou', 'Amber', 'Colette', 'Surge', 'Sprout', 'Nani', 'Gale', 'Jacky', 'Max', 'Mr.P',
        'Emz', 'Bea', 'Sandy', '8-Bit', 'Bibi', 'Carl', 'Rosa', 'Leon', 'Tick', 'Gene', 'Frank', 'Penny', 'Darryl',
        'Tara', 'Pam', 'Piper', 'Bo', 'Poco', 'Crow', 'Mortis', 'El Primo', 'Dynamike', 'Nita', 'Jessie', 'Barley',
        'Spike', 'Rico', 'Brock', 'Bull', 'Colt', 'Shelly'
    ]

    # Mapeamento entre nomes dos Brawlers em maiúsculas e caminhos das imagens
    imagens = {}

    # Gerar o mapeamento
    for nome in nomes_brawlers:
        # Converter o nome para maiúsculas
        nome_maiusculo = nome.upper()
        # Substituir espaços por underline para criar o nome do arquivo
        nome_arquivo = nome_maiusculo.replace(' ', '_')
        # Adicionar ao mapeamento
        imagens[nome_maiusculo] = os.path.join(diretorio_imagens, f'{nome_arquivo}_portrait.png')

    # Exibir o mapeamento gerado
    print(imagens)

    # Percorrer as células da planilha
    for row in workbook.iter_rows():
        for cell in row:
            # Verificar se o valor da célula é um nome com correspondência no dicionário de imagens
            if cell.value in imagens:
                # Carregar a imagem correspondente
                img = Image(imagens[cell.value])
                # Adicionar a imagem como uma forma na célula atual
                workbook.add_image(img, cell.coordinate)

    
    while True:
        try:
            workbook.save(f'{nomeArquivo}.xlsx')
            print("Documento formatado com sucesso.")
            break
        except PermissionError:
            input("Feche o arquivo para a formatação ser concluída. ")






# Nome do personagem que você quer procurar
nome_personagem = 'Tara'

# Convertendo o nome do personagem para minúsculas e substituindo espaços por underscores
nome_arquivo_imagem = nome_personagem.lower() + '_portrait.png'

# Caminho para a pasta contendo as imagens dos personagens
pasta_imagens = './brawlers/'

# Verificando se o arquivo de imagem correspondente ao nome do personagem existe na pasta
caminho_imagem = os.path.join(pasta_imagens, nome_arquivo_imagem)
if os.path.exists(caminho_imagem):
    print(f'O caminho da imagem para o personagem "{nome_personagem}" é: {caminho_imagem}')
else:
    print(f'A imagem para o personagem "{nome_personagem}" não foi encontrada na pasta.')
